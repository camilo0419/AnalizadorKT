# -*- coding: utf-8 -*-
import os
import re
import threading
import logging
import unicodedata
import pandas as pd

from tkinter import filedialog, Tk, Button, Label, messagebox, Entry, Toplevel
from tkinter import ttk

# Import robusto de pdfminer (igual Transporte)
try:
    from pdfminer.high_level import extract_text
except Exception:
    from pdfminer_high_level import extract_text  # type: ignore

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
import signal
import sys  # PyInstaller

# =========================
# Configuración general (IGUAL TRANSPORTE)
# =========================
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
    BASE_DIR = sys._MEIPASS
    RUNTIME_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    RUNTIME_DIR = BASE_DIR

# Ruta del Excel base (prioriza el archivo junto al .exe)
EXCEL_BASE = os.path.join(RUNTIME_DIR, "base_clausulas.xlsx")
if not os.path.isfile(EXCEL_BASE):
    EXCEL_BASE = os.path.join(BASE_DIR, "base_clausulas.xlsx")

# Silenciar mensajes ruidosos de pdfminer
for name in ["pdfminer", "pdfminer.layout", "pdfminer.converter", "pdfminer.image", "pdfminer.pdfinterp"]:
    logging.getLogger(name).setLevel(logging.ERROR)

# =========================
# Normalización y utilidades (IGUAL TRANSPORTE)
# =========================
PAGE_BREAK = "\f"  # \x0c

def strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def normalize_spaces_keep_newlines(s: str) -> str:
    s = re.sub(r'[ \t\r\f\v]+', ' ', s)
    s = re.sub(r' ?\n ?', '\n', s)
    return s

def fix_hyphen_linebreaks(s: str) -> str:
    return re.sub(r'-\s*\n\s*', '', s)

def fix_spaced_caps(s: str) -> str:
    def join_spaced_caps(m):
        return m.group(0).replace(' ', '')
    pattern = r'(?:(?<=\s)|^)(?:[A-ZÁÉÍÓÚÑ]\s){3,}[A-ZÁÉÍÓÚÑ](?=\s|[,.;:()\-\n]|$)'
    return re.sub(pattern, join_spaced_caps, s)

def repair_upper_sequences(pdf_text: str) -> str:
    parts = pdf_text.split(PAGE_BREAK)
    repaired_pages = []
    for part in parts:
        lines = part.splitlines()
        out, i = [], 0
        while i < len(lines):
            line = lines[i]
            if line.strip().isupper() and 1 <= len(line.strip()) <= 32:
                j, seq = i, []
                while j < len(lines) and (lines[j].strip().isupper() or lines[j].strip() == '') and len(seq) < 6:
                    if lines[j].strip():
                        seq.append(lines[j].strip())
                    j += 1
                if len(seq) >= 2:
                    out.append(' '.join(seq)); i = j; continue
            out.append(line); i += 1
        repaired_pages.append('\n'.join(out))
    return PAGE_BREAK.join(repaired_pages)

def normalize_pdf_text(pdf_path: str):
    raw = extract_text(pdf_path)
    step1 = fix_hyphen_linebreaks(raw)
    step2 = repair_upper_sequences(step1)
    step3 = fix_spaced_caps(step2)
    repaired_text = step3
    norm_lines = strip_accents(repaired_text.lower())
    norm_lines = normalize_spaces_keep_newlines(norm_lines)
    norm_compact = re.sub(r'\s+', ' ', norm_lines.replace(PAGE_BREAK, ' ')).strip()
    return repaired_text, norm_lines, norm_compact

def _cut_after_line(text: str, pos_end: int) -> str:
    line_end = text.find('\n', pos_end)
    if line_end == -1:
        line_end = len(text)
    cut_pos = min(line_end + 1, len(text))
    return text[cut_pos:]

def crop_from_cp_and_clausulas(repaired_text: str):
    raw = repaired_text
    low = strip_accents(raw.lower())

    m_cp = re.search(r'^\s*condiciones\s+particulares\s*:?\s*$', low, flags=re.M)
    if m_cp:
        raw = _cut_after_line(raw, m_cp.end())
        low = strip_accents(raw.lower())

    m_cl = re.search(r'^\s*cl[aá]usulas\s*:?\s*$', low, flags=re.M)
    if m_cl:
        raw = _cut_after_line(raw, m_cl.end())

    norm_lines = strip_accents(raw.lower())
    norm_lines = normalize_spaces_keep_newlines(norm_lines)
    norm_compact = re.sub(r'\s+', ' ', norm_lines.replace(PAGE_BREAK, ' ')).strip()
    return raw, norm_lines, norm_compact

def normalized_basic(s: str) -> str:
    s = strip_accents(s.lower())
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


# =========================
# Tokens útiles (SOLO MRC) para evitar falsos positivos
# - No contamos palabras genéricas (ej. "amparo") para el Jaccard,
#   porque disparan coincidencias falsas como "Amparo animales vivos".
# =========================
GENERIC_TOKENS = {
    "amparo", "clausula", "cláusula", "clausulas", "cláusulas", "cobertura",
    "condicion", "condiciones", "general", "generales", "particular", "particulares",
    "anexo", "anexos", "seccion", "secciones"
}

STRICT_EXACT_TITLES = {
    # Evita falsos positivos: SOLO se marca si existe coincidencia EXACTA
    "amparo animales vivos",
}

def tokens_utiles(s: str) -> set:
    toks = normalized_basic(s).split()
    return {t for t in toks if t not in GENERIC_TOKENS}

# =========================
# Detección de CUERPO / títulos (IGUAL TRANSPORTE)
# =========================
def is_body_line(s: str) -> bool:
    T = s.strip().lower()
    if re.match(r'^\s*\d+\)\s', T):
        return True
    if T.endswith('.') and len(T.split()) >= 18:
        return True
    return False

def is_title_like(norm_text: str) -> bool:
    T = norm_text.strip()
    if len(T) < 4 or len(T) > 320:
        return False
    if re.match(r'^\d+\)\s', T):
        return False
    if T.endswith('.') and len(T.split()) >= 18:
        return False
    return True

# =========================
# Candidatos (combos dentro de página)
# =========================
def build_title_candidates(norm_lines: str, repaired_text: str, max_lines_combo=5):
    rep_lines = repaired_text.split('\n')
    norm_lns  = norm_lines.split('\n')

    # Transporte usaba assert; aquí lo hacemos tolerante
    if len(rep_lines) != len(norm_lns):
        N = min(len(rep_lines), len(norm_lns))
        rep_lines = rep_lines[:N]
        norm_lns  = norm_lns[:N]

    line_starts, acc = [], 0
    for s in rep_lines:
        line_starts.append(acc); acc += len(s) + 1

    cand = []
    N = len(rep_lines)

    for i in range(N):
        base_norm = norm_lns[i].strip()
        if not base_norm:
            continue
        if is_body_line(rep_lines[i]):
            continue

        for k in range(1, max_lines_combo + 1):
            j = i + k - 1
            if j >= N:
                break
            if k > 1 and any(is_body_line(rep_lines[t]) for t in range(i + 1, j + 1)):
                break

            combo_norm = ' '.join(norm_lns[i:j+1]).strip()
            if not is_title_like(combo_norm):
                continue

            combo_orig = ' '.join(rep_lines[i:j+1]).strip()
            if combo_orig.strip().endswith('.') and len(combo_orig.split()) >= 18:
                continue

            cand.append({
                'text_basic': normalized_basic(combo_norm),
                'pos': line_starts[i],
                'orig': re.sub(r'\s+', ' ', combo_orig)
            })

    seen, uniq = set(), []
    for c in cand:
        key = (c['text_basic'], c['pos'])
        if key not in seen:
            seen.add(key); uniq.append(c)

    uniq.sort(key=lambda x: x['pos'])
    return uniq

def jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b); union = len(a | b)
    return inter / union if union else 0.0

# =========================
# Matching (IGUAL TRANSPORTE)
# =========================
def match_titles_against_candidates(titles: list, candidates: list):
    results = []
    used_candidates = set()

    titles_map = {}
    for idx, title in enumerate(titles):
        t_key = normalized_basic(title)
        if t_key:
            titles_map.setdefault(t_key, []).append(idx)

    found_map = {}

    # Exactos
    for c_idx, c in enumerate(candidates):
        key = c['text_basic']
        if key in titles_map and c_idx not in used_candidates:
            for t_idx in titles_map[key]:
                if t_idx not in found_map:
                    found_map[t_idx] = {'pos': c['pos'], 'obs': "Exacta", 'used_idx': c_idx}
                    used_candidates.add(c_idx)
                    break

    # Robustos (Jaccard 0.60)
    for t_idx, title in enumerate(titles):
        if t_idx in found_map:
            continue
        t_key = normalized_basic(title)
        if t_key in STRICT_EXACT_TITLES:
            continue

        t_tokens = tokens_utiles(title)
        if not t_tokens:
            t_tokens = set(normalized_basic(title).split())
        best = None
        for c_idx, c in enumerate(candidates):
            if c_idx in used_candidates:
                continue
            # tokens útiles: usamos el texto original del candidato para no perder palabras
            c_tokens = tokens_utiles(c.get('orig', c['text_basic']))
            if not c_tokens:
                c_tokens = set(c['text_basic'].split())
            jac = jaccard(t_tokens, c_tokens)
            if jac >= 0.60:
                score = jac
                if (best is None) or (score > best[0]) or (score == best[0] and c['pos'] < best[2]['pos']):
                    best = (score, c_idx, c)
        if best is not None:
            _, c_idx, c = best
            found_map[t_idx] = {'pos': c['pos'], 'obs': "Robusta", 'used_idx': c_idx}
            used_candidates.add(c_idx)

    for idx, _ in enumerate(titles):
        if idx in found_map:
            r = found_map[idx]
            results.append((True, r['pos'], r['obs']))
        else:
            results.append((False, float('inf'), "No"))
    return results

def fallback_footer_titles(titles, repaired_text, already_matched, last_k_lines=8):
    found = {}
    pages = repaired_text.split(PAGE_BREAK)
    pos_global = 0

    def nbasic(s):
        return normalized_basic(s)

    for page in pages:
        rep_lines = page.split('\n')

        line_starts, acc = [], 0
        for s in rep_lines:
            line_starts.append(acc)
            acc += len(s) + 1

        non_empty = [i for i, ln in enumerate(rep_lines) if ln.strip()]
        footer_idx = non_empty[-last_k_lines:] if non_empty else []

        for i in footer_idx:
            line = rep_lines[i].strip()
            if not line or is_body_line(line):
                continue

            c_tokens = tokens_utiles(line)
            if not c_tokens:
                c_tokens = set(nbasic(line).split())

            for t_idx, title in enumerate(titles):
                # Títulos "estrictos": NO pasan por footer (solo exactos)
                t_key = nbasic(title)
                if t_key in STRICT_EXACT_TITLES:
                    continue

                if already_matched[t_idx] or t_idx in found:
                    continue

                t_tokens = tokens_utiles(title)
                if not t_tokens:
                    t_tokens = set(nbasic(title).split())

                if jaccard(t_tokens, c_tokens) >= 0.60:
                    found[t_idx] = (pos_global + line_starts[i], "Footer")

        pos_global += sum(len(s) + 1 for s in rep_lines) + 1

    return found

# =========================
# Extra MRC: Valor asegurado (PDF) (opcional)
# =========================
_MONEY_RE = re.compile(
    r'(?:(?:cop|col)\s*)?\$?\s*\d{1,3}(?:\.\d{3})+(?:,\d{2})?'
    r'|\$?\s*\d{1,3}(?:,\d{3})+(?:\.\d{2})?'
    r'|\b\d{1,3}\s*%\b'
    r'|\b\d+\s*(?:smmlv|smdlv)\b',
    flags=re.IGNORECASE
)

def extraer_valor_asegurado_pdf(repaired_text: str, pos: int, window: int = 1400) -> str:
    if pos is None or pos == float('inf'):
        return ""
    chunk = repaired_text[pos: pos + window]
    chunk_low = strip_accents(chunk.lower())
    chunk_low = re.sub(r'pagina\s+\d+\s+de\s+\d+', ' ', chunk_low)
    m = _MONEY_RE.search(chunk_low)
    if not m:
        return ""
    # devolvemos el match sobre chunk_low, normalizando espacios
    return re.sub(r'\s+', ' ', m.group(0)).strip()

# =========================
# Motor principal (MRC) - con mismo "Índice de Orden" real del PDF
# =========================
def extraer_clausulas_por_titulo_mrc(pdf_path, excel_base, progress_bar, root):
    if not os.path.exists(excel_base):
        messagebox.showerror("Error", f"No se encuentra el archivo:\n{excel_base}")
        return [], 0, 0

    try:
        df = pd.read_excel(excel_base)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo leer el Excel:\n{e}")
        return [], 0, 0

    # Columna de título MRC (obligatoria)
    col_titulo = "Multiriesgo Corporativo" if "Multiriesgo Corporativo" in df.columns else None
    if col_titulo is None or "Texto de la cláusula" not in df.columns:
        messagebox.showerror(
            "Error",
            "El Excel base debe tener:\n- 'Multiriesgo Corporativo'\n- 'Texto de la cláusula'"
        )
        return [], 0, 0

    titles = df[col_titulo].fillna("").astype(str).tolist()
    texts  = df["Texto de la cláusula"].fillna("").astype(str).tolist()

    # Columnas extra del base (se arrastran tal cual)
    base_tipo  = df["Tipo de operación"].fillna("").astype(str).tolist() if "Tipo de operación" in df.columns else [""] * len(titles)
    base_valor = df["Valor asegurado"].fillna("").astype(str).tolist() if "Valor asegurado" in df.columns else [""] * len(titles)
    base_obs   = df["Observaciones"].fillna("").astype(str).tolist() if "Observaciones" in df.columns else [""] * len(titles)
    base_lucro = df["SOLO PARA LUCRO"].fillna("").astype(str).tolist() if "SOLO PARA LUCRO" in df.columns else [""] * len(titles)
    base_aclar = df["ACLARACIONES "].fillna("").astype(str).tolist() if "ACLARACIONES " in df.columns else [""] * len(titles)

    # PDF
    try:
        repaired_text, _, _ = normalize_pdf_text(pdf_path)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo leer el PDF:\n{e}")
        return [], 0, 0

    repaired_text, norm_lines, _ = crop_from_cp_and_clausulas(repaired_text)

    candidates = build_title_candidates(norm_lines, repaired_text, max_lines_combo=5)
    matches = match_titles_against_candidates(titles, candidates)

    already = [f for (f, _, _) in matches]
    footer_hits = fallback_footer_titles(titles, repaired_text, already, last_k_lines=8)

    final_matches = []
    for idx, (f, p, o) in enumerate(matches):
        if not f and idx in footer_hits:
            pos, obs = footer_hits[idx]
            final_matches.append((True, pos, obs))
        else:
            final_matches.append((f, p, o))

    # Índice de Orden real por aparición (IGUAL Transporte)
    found_positions = [(i, m[1]) for i, m in enumerate(final_matches) if m[0]]
    found_positions.sort(key=lambda x: x[1])
    rank_by_tidx = {t_idx: rank for rank, (t_idx, _) in enumerate(found_positions, start=1)}

    resultados = []
    found_total = 0

    for idx, ((found, pos, obs), title, txt) in enumerate(zip(final_matches, titles, texts)):
        if found:
            found_total += 1
            indice = rank_by_tidx.get(idx, "N/A")
            valor_pdf = extraer_valor_asegurado_pdf(repaired_text, pos)
        else:
            indice = "N/A"
            pos = float('inf')
            obs = ""
            valor_pdf = ""

        resultados.append({
            "Indice de Orden": indice,
            "Multiriesgo Corporativo": title,
            "Texto de la cláusula": txt,
            "Encontrado": "Sí" if found else "No",
            "Tipo de operación": "",
            "Valor asegurado": base_valor[idx] if idx < len(base_valor) else "",
            "Observaciones": base_obs[idx] if idx < len(base_obs) else "",
            "Compatibilidad": obs,  # Exacta / Robusta / Footer
            "Posicion": pos,
            "OrdenBase": idx
        })

    resultados.sort(key=lambda x: (x["Encontrado"] != "Sí", x["Posicion"], x["OrdenBase"]))
    return resultados, len(titles), found_total

# =========================
# Guardado en Excel (estilo transporte)
# =========================
def guardar_resultados_en_excel(resultados, nombre_salida):
    df_salida = pd.DataFrame(resultados)

    columnas_finales = [
        "Indice de Orden",
        "Multiriesgo Corporativo",
        "Texto de la cláusula",
        "Tipo de operación",
        "Valor asegurado",
        "Observaciones",
        "Encontrado",
        "Compatibilidad"
    ]
    # Compat: en el base había "ACLARACIONES " con espacio; aquí ya lo normalizamos a "ACLARACIONES"
    df_salida = df_salida[columnas_finales]
    df_salida.to_excel(nombre_salida, index=False)

    wb = load_workbook(nombre_salida)
    ws = wb.active

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName="Resultados", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=False, showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    font_header = Font(bold=True, color="FFFFFF", name="Calibri", size=8)
    fill_header = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    font_body = Font(name="Calibri", size=8)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    widths = [15, 40, 105, 22, 20, 30, 12, 18]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = font_body
            cell.alignment = align_center
        ws.row_dimensions[row[0].row].height = 60
    # Lista desplegable para "Tipo de operación" (sin depender de separadores regionales)
    # Creamos una hoja oculta con las opciones y referenciamos el rango.
    try:
        if "Listas" in wb.sheetnames:
            ws_listas = wb["Listas"]
            # limpiar rango A1:A5
            for r in range(1, 6):
                ws_listas.cell(row=r, column=1).value = None
        else:
            ws_listas = wb.create_sheet("Listas")

        opciones = ["Ajustar", "Cubre", "Incluir", "No aplica", "Retirar"]
        for i, val in enumerate(opciones, start=1):
            ws_listas.cell(row=i, column=1, value=val)

        # Ocultar hoja
        ws_listas.sheet_state = "hidden"

        dv = DataValidation(type="list", formula1="=Listas!$A$1:$A$5", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"D2:D{ws.max_row}")
    except Exception:
        pass

    wb.save(nombre_salida)
    return nombre_salida

# =========================
# Hilo y GUI (igual transporte)
# =========================
def run_analysis_thread(ruta_pdf, progress_bar, root):
    try:
        root.after(0, lambda: progress_bar.config(mode="indeterminate", style="blue.Horizontal.TProgressbar"))
        root.after(0, lambda: progress_bar.start())

        resultados, total_clausulas, encontradas_count = extraer_clausulas_por_titulo_mrc(
            ruta_pdf, EXCEL_BASE, progress_bar, root
        )

        root.after(0, lambda: progress_bar.stop())
        root.after(0, lambda: progress_bar.config(mode="determinate", value=100, style="blue.Horizontal.TProgressbar"))

        if not resultados:
            root.after(100, lambda: messagebox.showwarning("Aviso", "No se generaron resultados."))
            return

        ruta_salida = elegir_ruta_guardado(ruta_pdf)
        if not ruta_salida:
            root.after(100, lambda: messagebox.showinfo("Cancelado", "Guardado cancelado por el usuario."))
            return

        try:
            guardar_resultados_en_excel(resultados, ruta_salida)
        except PermissionError:
            msg = ("No se pudo guardar el archivo.\n\n"
                   "Es posible que el archivo de destino esté ABIERTO o protegido.\n"
                   "Ciérralo o elige otra ruta y vuelve a intentarlo.")
            root.after(100, lambda m=msg: messagebox.showerror("Error de permisos", m))
            return
        except Exception as err:
            msg = f"Ocurrió un error al guardar:\n{err}"
            root.after(100, lambda m=msg: messagebox.showerror("Error", m))
            return

        no_encontradas_count = total_clausulas - encontradas_count
        porcentaje = (encontradas_count / total_clausulas) * 100 if total_clausulas > 0 else 0

        mensaje = (
            "¡Análisis completado!\n"
            f"Se guardó en:\n{ruta_salida}\n\n"
            "Estadísticas:\n"
            f"Total de cláusulas: {total_clausulas}\n"
            f"Encontradas: {encontradas_count} ({porcentaje:.1f}%)\n"
            f"No encontradas: {no_encontradas_count} ({100-porcentaje:.1f}%)"
        )
        root.after(100, lambda m=mensaje: messagebox.showinfo("¡Listo!", m))

    except Exception as err:
        msg = f"Ocurrió un error:\n{err}"
        root.after(100, lambda m=msg: messagebox.showerror("Error", m))
    finally:
        root.after(100, lambda: progress_bar.pack_forget())

def elegir_ruta_guardado(ruta_pdf: str) -> str | None:
    carpeta_pdf = os.path.dirname(ruta_pdf)
    pdf_nombre = os.path.basename(ruta_pdf)
    # Mismo patrón que Transporte (quita prefijo 00_Sura_)
    nombre_sin_prefijo = re.sub(r'^\d{2}_(Sura|SURA)_', '', pdf_nombre)
    nombre_sugerido = os.path.splitext(nombre_sin_prefijo)[0] + "_MRC.xlsx"

    ruta_destino = filedialog.asksaveasfilename(
        title="Guardar resultados como...",
        initialdir=carpeta_pdf,
        initialfile=nombre_sugerido,
        defaultextension=".xlsx",
        filetypes=[("Archivo de Excel", "*.xlsx")],
        confirmoverwrite=True
    )
    return ruta_destino if ruta_destino else None

def seleccionar_pdf_y_procesar():
    ruta_pdf = filedialog.askopenfilename(
        title="Selecciona un archivo PDF",
        filetypes=[("Archivos PDF", "*.pdf")]
    )
    if ruta_pdf:
        progress_bar.pack(pady=10)
        root.update_idletasks()
        t = threading.Thread(target=run_analysis_thread, args=(ruta_pdf, progress_bar, root))
        t.daemon = True
        t.start()

def mostrar_login(root, on_success):
    login_win = Toplevel(root)
    login_win.title("Acceso")
    login_win.geometry("300x180")
    login_win.configure(bg="#0070C0")
    login_win.grab_set()

    # Intentar cargar el ícono (prioriza el que está junto al .exe)
    try:
        ico_path = os.path.join(RUNTIME_DIR, "icono.ico")
        if not os.path.isfile(ico_path):
            ico_path = os.path.join(BASE_DIR, "icono.ico")
        login_win.iconbitmap(ico_path)
    except Exception:
        pass

    Label(login_win, text="Ingrese la contraseña", font=("Arial", 12, "bold"),
          bg="#0070C0", fg="white").pack(pady=15)

    entry = Entry(login_win, show="*", font=("Arial", 12))
    entry.pack(pady=10)
    entry.focus()

    def verificar():
        if entry.get() == "kt1324":
            login_win.destroy()
            on_success()
        else:
            messagebox.showerror("Error", "Contraseña incorrecta")
            entry.delete(0, "end")

    Button(login_win, text="Ingresar", command=verificar,
           bg="white", fg="#0070C0", font=("Arial", 11, "bold"),
           padx=10, pady=5).pack(pady=10)

    login_win.bind("<Return>", lambda e: verificar())

def construir_app(root):
    global progress_bar
    root.title("Analizador de Cláusulas – MRC")
    try:
        root.iconbitmap(os.path.join(BASE_DIR, 'icono.ico'))
    except Exception:
        pass
    root.geometry("860x420")

    Label(root, text="Analizador de Cláusulas – MRC", font=("Arial", 16, "bold")).pack(pady=10)
    Label(root, text="Reporte ordenado de cláusulas incluidas en el documento PDF.", font=("Arial", 11)).pack(pady=2)
    Label(root, text="Matching Exacto/Robusto + Fallback Footer", font=("Arial", 10)).pack(pady=2)
    Label(root, text="Selecciona un archivo PDF - MRC para analizar", font=("Arial", 14)).pack(pady=20)

    Button(root, text="Seleccionar PDF", command=seleccionar_pdf_y_procesar,
           font=("Arial", 12), bg="#0070C0", fg="white", padx=20, pady=10).pack(pady=10)

    style = ttk.Style()
    style.theme_use('default')
    style.configure("blue.Horizontal.TProgressbar", background='#0070C0', troughcolor='#e0e0e0')

    progress_bar = ttk.Progressbar(root, orient="horizontal", length=320,
                                   mode="determinate", style="blue.Horizontal.TProgressbar")

    Label(root, text="Nota: 'base_clausulas.xlsx' debe estar en la misma carpeta donde se ejecuta el programa.", font=("Arial", 10), fg="gray").pack(pady=(10, 0))
    root.protocol("WM_DELETE_WINDOW", root.quit)

# -------- INICIO --------
if __name__ == "__main__":
    root = Tk()
    root.withdraw()

    def on_login_ok():
        root.deiconify()
        construir_app(root)

    mostrar_login(root, on_login_ok)

    try:
        signal.signal(signal.SIGINT, lambda *args: root.quit())
    except Exception:
        pass

    try:
        root.mainloop()
    except KeyboardInterrupt:
        pass
